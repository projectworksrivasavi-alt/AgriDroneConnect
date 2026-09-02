from openpyxl import Workbook, load_workbook
import os

FILE_NAME = "AgriData.xlsx"

def create_excel():

    if not os.path.exists(FILE_NAME):

        wb = Workbook()

        # Farmers Sheet
        ws1 = wb.active
        ws1.title = "Farmers"
        ws1.append([
            "Farmer ID",
            "Name",
            "Phone",
            "Village",
            "District",
            "State"
        ])

        # Operators Sheet
        ws2 = wb.create_sheet("Operators")
        ws2.append([
            "Operator ID",
            "Name",
            "Phone",
            "Drone Type",
            "Village",
            "District",
            "State"
        ])

        # Bookings Sheet
        ws3 = wb.create_sheet("Bookings")
        ws3.append([
            "Booking ID",
            "Farmer Name",
            "Operator Name",
            "Service",
            "Date",
            "Status"
        ])

        wb.save(FILE_NAME)


def save_farmer(data):

    create_excel()

    wb = load_workbook(FILE_NAME)
    ws = wb["Farmers"]

    ws.append([
        str(data.get("_id")),
        data.get("name"),
        data.get("phone"),
        data.get("village"),
        data.get("district"),
        data.get("state")
    ])

    wb.save(FILE_NAME)


def save_operator(data):

    create_excel()

    wb = load_workbook(FILE_NAME)
    ws = wb["Operators"]

    ws.append([
        str(data.get("_id")),
        data.get("name"),
        data.get("phone"),
        data.get("drone_type"),
        data.get("village"),
        data.get("district"),
        data.get("state")
    ])

    wb.save(FILE_NAME)


def save_booking(data):

    create_excel()

    wb = load_workbook(FILE_NAME)
    ws = wb["Bookings"]

    ws.append([
        str(data.get("_id")),
        data.get("farmer_name"),
        data.get("operator_name"),
        data.get("service"),
        data.get("date"),
        data.get("status")
    ])

    wb.save(FILE_NAME)